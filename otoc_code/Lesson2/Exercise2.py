# 练习
# 湖人队的常规赛和季前赛的数据
# 1.删除所有季前赛的数据行
# 2.在第一行插入19-20常规赛数据
# ["19-20","常规赛",113.4,48,88.3,34.9,31.6,72.9,24.3,45.7,25.4,8.6,6.6,20.7]
# 3.把所有的数据往下移一行，第一行的内容是
# ["赛季","","得分","命中%","出手","三分%","三分出手","罚球%","罚球出手","篮板","助攻","抢断","盖帽","犯规"]
# 4.合并A1和B1

from openpyxl import load_workbook
fileName = "2.xlsx"
wb = load_workbook(fileName)
ws = wb["nba"]
datas = ["19-20","常规赛",113.4,48,88.3,34.9,31.6,72.9,24.3,45.7,25.4,8.6,6.6,20.7]
titles = ["赛季","","得分","命中%","出手","三分%","三分出手","罚球%","罚球出手","篮板","助攻","抢断","盖帽","犯规"]
deletRows = []
for cell in ws["B"]:
    if cell.value == '季前赛':
        deletRows.append(cell.row)
for index in range(len(deletRows)-1,-1,-1):
    ws.delete_rows(deletRows[index])

ws.insert_rows(1)
index = 0
for row in ws.iter_rows(min_row=1,max_row=1,min_col=1,max_col=14):
    for cell in row:
        cell.value = datas[index]
        index += 1

ws.move_range("A1:N4",rows=1)

index = 0
for row in ws.iter_rows(min_row=1,max_row=1,min_col=1,max_col=14):
    for cell in row:
        cell.value = titles[index]
        index += 1
ws.merge_cells("A1:B1")

wb.save(fileName)