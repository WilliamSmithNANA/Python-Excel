from openpyxl import load_workbook
from openpyxl.drawing.image import Image
# import random
# from GenerateExcel import random_name
# from datetime import datetime
# from datetime import timedelta
# Section1 插入、删除行列、移动区域


fileName = "2.xlsx"
wb = load_workbook(fileName)
ws = wb.active
# 插入一行
ws.insert_rows(6)
row6 = ws[6]
newDatas = ["余小飞","男","1994-03-22",26]
for row in ws["A6":"D6"]:
    for index in range(len(row)):
        row[index].value = newDatas[index]

# 删除多行
ws.delete_rows(7,2)

# Section2 移动整块区域
ws.move_range("A1:D12",rows=4,cols=2)
ws.move_range("A1:D12",rows=-1,cols=-2)

# Section2 插入图片和合并拆分单元格
# 插入图片,合并单元格，必需先安装Pillow
ws1 = wb.create_sheet("beauty")
ws1.merge_cells("A1:G27")
ws1.merge_cells("I1:Q47")
# unmerge_cells
ws1.add_image(img1,"A1")
ws1.add_image(img2,"I1")

# Section3 巩固练习
# 湖人队的常规赛和季前赛的数据
# 1.删除所有季前赛的数据行
# 2.在第一行插入19-20常规赛数据
# ["19-20","常规赛",113.4,48,88.3,34.9,31.6,72.9,24.3,45.7,25.4,8.6,6.6,20.7]
# 3.把所有的数据往下移一行，第一行的内容是
# ["赛季","","得分","命中%","出手","三分%","三分出手","罚球%","罚球出手","篮板","助攻","抢断","盖帽","犯规"]
# 4.合并A1和B1

ws = wb["nba"]
datas = ["19-20","常规赛",113.4,48,88.3,34.9,31.6,72.9,24.3,45.7,25.4,8.6,6.6,20.7]
titles = ["赛季","","得分","命中%","出手","三分%","三分出手","罚球%","罚球出手","篮板","助攻","抢断","盖帽","犯规"]
deletRows = []
for cell in ws["B"]:
    if cell.value == '季前赛':
        deletRows.append(cell.row)
for index in range(len(deletRows)-1,-1,-1):
    ws.delete_rows(deletRows[index])

ws.insert_rows(1)
index = 0
for row in ws.iter_rows(min_row=1,max_row=1,min_col=1,max_col=14):
    for cell in row:
        cell.value = datas[index]
        index += 1

ws.move_range("A1:N4",rows=1)

index = 0
for row in ws.iter_rows(min_row=1,max_row=1,min_col=1,max_col=14):
    for cell in row:
        cell.value = titles[index]
        index += 1
ws.merge_cells("A1:B1")

wb.save(fileName)


