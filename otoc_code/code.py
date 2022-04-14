import os#提供通用的、基本的操作系统的交互功能
#from import导导入模块中指定的属性
from openpyxl import Workbook#用来操作excel文件
from openpyxl import load_workbook#用来加载已经存在的excel文件
#.的作用跟上面的from import是一样的，这里我们从openpyxl里面的styles里面在导入alignment，font
from openpyxl.styles import Alignment,Font#alignment的作用是用来设置对齐方式的，font的作用是用来设置字体格式的

#Part1
#获得当前的绝对路径
print(os.path.abspath('.'))
#将当前文件夹下的所有excel文件的文件名都存入到xlfs这个列表当中
xlfs = [
    #os.listdir(path)————列出path目录下所有的文件和目录名
    #os.path.isfile(path)————判断指定对象是否是文件，是返回true，否返回false
    #os.path.splitext(path)————将指定对象的扩展名分离
    x for x in os.listdir('.') if os.path.isfile(x)
    and os.path.splitext(x)[1]=='.xlsx'
]
print("需要统计",len(xlfs),"个excel文件")
print(xlfs)

#Part2 
#给起个文件名
fileName = '签到表.xlsx'
#创建一个excel文件
wb = Workbook()
#打开这个excel文件
ws = wb.active
#编辑这个excel文件
#制作表头
ws['A1'] = '签到表'
ws.merge_cells('A1:H1')
ws['A1'].font = Font(name='微软雅黑',size=18,bold=True)
ws['A1'].alignment = Alignment(horizontal='center',vertical='center')#horizontal是水平，vertical是竖直
lis = ['序号','姓名','性别','单位简称','岗位/职务','手机号码','身份证号','签到']
i = 2
j = 1
for sheetTitle in lis:
    ws.cell(row=i, column=j, value=sheetTitle)
    j = j+1
#汇总回执中的内容到签到表中
#利用循环读取文件名
x = 3
num = 1
for wb_name in xlfs:
    #加载当前的excel文件
    wb_temp = load_workbook(wb_name)
    #打开当前加载的excel文件
    ws_temp = wb_temp.active
    #按行读取当前的excel文件，从第5行第2列开始读取
    row_list = ws_temp.iter_rows(min_row=5,min_col=2)
    #先读取临时变量中一行的数据
    for row in row_list:
        #把读取的一行数据输入到生成的签到表中
        ws.cell(row=x, column=1,value=num)
        y = 2
        for temp in row:
            ws.cell(row=x, column=y, value=temp.value)
            y = y+1
        x = x+1
        num = num+1
#制作表尾
ws.cell(row=x, column=1,value='班主任签字')
ws.merge_cells(start_row=x,start_column=1,end_row=x,end_column=4)
ws.cell(row=x, column=1).font = Font(name='微软雅黑',size=11,bold=True)
ws.cell(row=x, column=5,value='教研室主任签字')
ws.merge_cells(start_row=x,start_column=5,end_row=x,end_column=8)
ws.cell(row=x, column=5).font = Font(name='微软雅黑',size=11,bold=True)
#保存这个excel文件
wb.save(fileName)
#关闭这个excel文件
wb.close()
