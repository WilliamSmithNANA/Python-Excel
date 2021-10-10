# 练习
# 在1.xlsx 新建一个名为 "Exercise" sheet,在["A1":"D6"]区域，随机赋值
# 找到这块区域中的最大值，最小值，分别赋值到"D7","D8",
# 求出这块区域的和，赋值给"D9"
from openpyxl import Workbook
from openpyxl import load_workbook
import random

fileName = "1.xlsx"
wb = load_workbook(fileName)
ws = wb.create_sheet("Exercise")
cell_range = ws["A1":"D6"]
for row in cell_range:
    for cell in row:
        cell.value = random.randint(1,100)


maxValue = 0
minValue = 101
sum = 0
for row in ws.iter_rows(min_row=1,max_row=6,min_col=1,max_col=4,values_only=True):
    for val in row:
        if val > maxValue:
            maxValue = val
        if val < minValue:
            minValue = val
        sum += val

ws["D7"] = maxValue
ws["D8"] = minValue
ws["D9"] = sum

wb.save(fileName)