from openpyxl import Workbook
from openpyxl import load_workbook
import random

# Section1 WorkBook与WorkSheet层面
# 创建excel表格
fileName = "1.xlsx"
wb = Workbook()
# 默认会创建一个名为"Sheet"的Sheet
ws = wb.active
print("默认创建的sheet名:",ws.title)
# 修改sheet的名称
ws.title = "Sheet1"
# 修改下面tab的颜色
ws.sheet_properties.tabColor = "FF0000"

# 创建新sheet,可以控制位置，都表示是在这个位置之前插入
ws1 = wb.create_sheet("MySheet1")
ws2 = wb.create_sheet("MySheet2",0)
ws3 = wb.create_sheet("MySheet3",-1)


# 遍历所有的sheet
for sheetname in wb.sheetnames:
    print("遍历sheet:",sheetname)

# 选中sheet
ws3 = wb["MySheet3"]
print("选中sheet name: ",ws3.title)
# 修改active sheet
wb.active = wb["MySheet3"]
ws3 = wb.active
print("当前active sheet: ",ws3.title)

# Section2 Cell层面
# 单元格赋值
ws3["A1"] = 1
# 获取单元格
cell = ws3["A1"]
print("A1的值: ",cell.value)

# 获取一块区域的所有单元格
cell_range = ws3["A1":"C5"]
for row in cell_range:
    for cell in row:
        print("行{0}列{1}值{2}".format(cell.row,cell.column,cell.value))
        cell.value = random.randint(1,100)


# iter_rows 遍历行
for row in ws3.iter_rows(min_row=1,min_col=2,max_row=2,max_col=3):
    for cell in row:
        print("ite_row 行{0}列{1}值{2}".format(cell.row,cell.column,cell.value))

# iter_cols

for col in ws3.iter_cols(min_row=1,min_col=2,max_row=2,max_col=3):
    for cell in col:
        print("ite_col 行{0}列{1}值{2}".format(cell.row, cell.column, cell.value))


# 只想值遍历
for row in ws3.values:
    print("值遍历:",row)
#  iter_rows和iter_cols也支持值遍历
for row in ws3.iter_rows(min_row=1,min_col=2,max_row=2,max_col=3,values_only=True):
    print("iter_row 值遍历: ",row)

# 保存
wb.save(fileName)


# 加载已存在的excel
wb = load_workbook(fileName)
ws = wb.active
for row in ws.values:
    print("已存在的excel: ",row)


# Section3 巩固练习
# 在1.xlsx 新建一个名为 "Exercise" sheet,在["A1":"D6"]区域，随机赋值
# 找到这块区域中的最大值，最小值，分别赋值到"D7","D8",
# 求出这块区域的和，赋值给"D9"

wb = load_workbook(fileName)
ws = wb.create_sheet("Exercise")
cell_range = ws["A1":"D6"]
for row in cell_range:
    for cell in row:
        cell.value = random.randint(1,100)


maxValue = 0
minValue = 101
sum = 0
for row in ws.iter_rows(min_row=1,max_row=6,min_col=1,max_col=4,values_only=True):
    for val in row:
        if val > maxValue:
            maxValue = val
        if val < minValue:
            minValue = val
        sum += val

ws["D7"] = maxValue
ws["D8"] = minValue
ws["D9"] = sum

wb.save(fileName)


